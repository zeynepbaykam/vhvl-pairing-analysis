import pandas as pd
import numpy as np
from sklearn.metrics.pairwise import cosine_similarity
import matplotlib.pyplot as plt
from matplotlib.patches import Patch
import os
import sys
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from arpeggio_analysis import build_feature_matrix
from scipy import stats
from scipy.stats import fisher_exact
import seaborn as sns
from statsmodels.stats.multitest import multipletests
from sklearn.cluster import DBSCAN
import matplotlib.lines as mlines
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import colorsys
from scipy.stats import chi2_contingency



## Analyses that involve features ##

def calculate_cosine_similarity(feature_matrix):
    results = []
    bispec_names = feature_matrix.index.str.split("_").str[0].unique()
    
    for bispec in bispec_names:
        try:
            arm1_cog = feature_matrix.loc[f"{bispec}_1_True"].values.reshape(1, -1)
            arm2_cog = feature_matrix.loc[f"{bispec}_2_True"].values.reshape(1, -1)
            arm1_noncog = feature_matrix.loc[f"{bispec}_1_False"].values.reshape(1, -1)
            arm2_noncog = feature_matrix.loc[f"{bispec}_2_False"].values.reshape(1, -1)
            
            results.append({
                "bispecific": bispec,
                "cog1_vs_cog2": cosine_similarity(arm1_cog, arm2_cog)[0][0],
                "noncog1_vs_noncog2": cosine_similarity(arm1_noncog, arm2_noncog)[0][0],
                "arm1_cog_vs_noncog": cosine_similarity(arm1_cog, arm1_noncog)[0][0],
                "arm2_cog_vs_noncog": cosine_similarity(arm2_cog, arm2_noncog)[0][0]
            })
        except KeyError:
            continue
    
    return pd.DataFrame(results)


def plot_cosine_similarities(cosine_df):
    fig, axes = plt.subplots(2, 2, figsize=(12, 10))
    
    comparisons = [
        ("cog1_vs_cog2", "Cognate Arm1 vs Cognate Arm2", axes[0, 0]),
        ("noncog1_vs_noncog2", "Non-cognate Arm1 vs Non-cognate Arm2", axes[0, 1]),
        ("arm1_cog_vs_noncog", "Arm1 Cognate vs Non-cognate", axes[1, 0]),
        ("arm2_cog_vs_noncog", "Arm2 Cognate vs Non-cognate", axes[1, 1])
    ]
    
    for col, title, ax in comparisons:
        ax.hist(cosine_df[col], bins=20, edgecolor='black')
        ax.set_title(title)
        ax.set_xlabel("Cosine Similarity")
        ax.set_ylabel("Frequency")
        ax.axvline(x=cosine_df[col].mean(), color='r', linestyle='--', 
                   label=f"Mean: {cosine_df[col].mean():.2f}")
        ax.legend()
    
    plt.tight_layout()
    plt.savefig("cosine_similarity_distributions.png")
    plt.close()


def run_statistical_tests(cosine_df):
    # Paired t-test: are cognate-vs-noncognate similarities higher than cog1-vs-cog2?
    t1, p1 = stats.ttest_rel(cosine_df["arm1_cog_vs_noncog"], cosine_df["cog1_vs_cog2"])
    t2, p2 = stats.ttest_rel(cosine_df["arm2_cog_vs_noncog"], cosine_df["cog1_vs_cog2"])
    
    # Paired t-test: are cog1_vs_cog2 and noncog1_vs_noncog2 different?
    t3, p3 = stats.ttest_rel(cosine_df["cog1_vs_cog2"], cosine_df["noncog1_vs_noncog2"])
    
    print(f"Arm1 cog vs noncog > cog1 vs cog2: t={t1:.3f}, p={p1:.4f}")
    print(f"Arm2 cog vs noncog > cog1 vs cog2: t={t2:.3f}, p={p2:.4f}")
    print(f"Cog1 vs cog2 vs noncog1 vs noncog2: t={t3:.3f}, p={p3:.4f}")


def compare_cognate_noncognate_features(feature_matrix):

    # Split into cognate and non-cognate
    cognate_mask = feature_matrix.index.str.split("_").str[-1] == "True"
    cognate_df = feature_matrix[cognate_mask]
    noncognate_df = feature_matrix[~cognate_mask]

    cognate_sorted = cognate_df.copy()
    cognate_sorted["bispecific"] = cognate_sorted.index.str.rsplit("_", n=2).str[0]
    cognate_sorted["arm"] = cognate_sorted.index.str.rsplit("_", n=2).str[1]
    cognate_sorted = cognate_sorted.sort_values(["bispecific", "arm"]).reset_index(drop=True)

    noncognate_sorted = noncognate_df.copy()
    noncognate_sorted["bispecific"] = noncognate_sorted.index.str.rsplit("_", n=2).str[0]
    noncognate_sorted["arm"] = noncognate_sorted.index.str.rsplit("_", n=2).str[1]
    noncognate_sorted = noncognate_sorted.sort_values(["bispecific", "arm"]).reset_index(drop=True)

    results = []
    for feature in feature_matrix.columns:  # use feature_matrix.columns not cognate_sorted.columns
        cog_vals = cognate_sorted[feature].values
        noncog_vals = noncognate_sorted[feature].values

        # McNemar's contingency table
        # a = both 1, b = cog 1 noncog 0, c = cog 0 noncog 1, d = both 0
        a = ((cog_vals == 1) & (noncog_vals == 1)).sum()
        b = ((cog_vals == 1) & (noncog_vals == 0)).sum()
        c = ((cog_vals == 0) & (noncog_vals == 1)).sum()
        d = ((cog_vals == 0) & (noncog_vals == 0)).sum()

        # McNemar's test - only discordant pairs b and c matter
        if b + c == 0:
            p_value = 1.0
        else:
            # Use exact binomial test for small samples, chi-squared for large
            if b + c < 25:
                result = stats.binomtest(b, b + c, p=0.5, alternative="two-sided")
                p_value = result.pvalue
            else:
                # McNemar's chi-squared with continuity correction
                chi2 = (abs(b - c) - 1) ** 2 / (b + c)
                p_value = stats.chi2.sf(chi2, df=1)

        results.append({
            "feature": feature,
            "cog_frequency": cog_vals.mean(),
            "noncog_frequency": noncog_vals.mean(),
            "frequency_diff": cog_vals.mean() - noncog_vals.mean(),
            "n_cog1_noncog0": b,
            "n_cog0_noncog1": c,
            "p_value": p_value
        })

    results_df = pd.DataFrame(results)
    _, results_df["p_adjusted"], _, _ = multipletests(results_df["p_value"], method="fdr_bh")
    results_df = results_df.sort_values("p_adjusted")
    return results_df

def plot_feature_differences(results_df, n_top=20):
    
    sig_features = results_df[results_df["p_adjusted"] < 0.05]
    nonsig_features = results_df[results_df["p_adjusted"] >= 0.05]
    
    if len(sig_features) == 0:
        print("No significant features found after multiple testing correction - plotting top features by frequency difference")
        plot_df = results_df.reindex(
            results_df["frequency_diff"].abs().sort_values(ascending=False).index
        ).head(n_top)
        title = f"Top {n_top} features by frequency difference (none significant after BH correction)"
    else:
        plot_df = results_df.sort_values("p_adjusted", ascending=True).head(n_top)
        title = "Top 10 features differing between cognate and non-cognate pairings"
    
    plot_df = plot_df.sort_values("p_adjusted", ascending=False)
    
    plt.figure(figsize=(12, 8))
    colors = ["#1A5276" if x > 0 else "#922B21" for x in plot_df["frequency_diff"]]
    plt.barh(plot_df["feature"], plot_df["frequency_diff"], color=colors)
    plt.axvline(x=0, color="black", linestyle="-")
    plt.xlabel("Frequency Difference (Cognate - Non-cognate)")
    plt.title(title)
    
    x_range = plot_df["frequency_diff"].abs().max()
    min_offset = x_range * 0.01
    
    for i, (_, row) in enumerate(plot_df.iterrows()):
        x_pos = row["frequency_diff"]
        is_sig = row["p_adjusted"] < 0.05
        p_text = f"p={row['p_adjusted']:.3f}*" if is_sig else f"p={row['p_adjusted']:.3f}"
        fontweight = "bold" if is_sig else "normal"
        
        if x_pos >= 0:
            plt.text(x_pos + min_offset, i, p_text,
                    va='center', ha='left', fontsize=7, fontweight=fontweight)
        else:
            plt.text(x_pos - min_offset, i, p_text,
                    va='center', ha='right', fontsize=7, fontweight=fontweight)
    
    legend_elements = [Patch(facecolor="#1A5276", label="More frequent in cognate"),
                      Patch(facecolor="#922B21", label="More frequent in non-cognate")]
    plt.legend(handles=legend_elements)
    plt.tight_layout()
    plt.savefig("feature_differences.png", dpi=150, bbox_inches='tight')
    plt.close()
    print("Saved feature_differences.png")

## Analysis on total counts per interaction type ##
def calculate_interaction_counts(bispec_interactions):
    interaction_cols = ['AMIDEAMIDE', 'AMIDERING', 'CARBONPI', 'CATIONPI', 'DONORPI', 
                       'EE', 'EF', 'ET', 'FE', 'FF', 'FT', 'METSULPHURPI', 'OE', 'OF', 
                       'OT', 'aromatic', 'carbonyl', 'covalent', 'hbond', 'hydrophobic', 
                       'ionic', 'polar', 'proximal', 'vdw', 'vdw_clash', 'weak_hbond', 
                       'weak_polar']
    
    counts_df = bispec_interactions.groupby("antibody")[interaction_cols].sum().reset_index()
    counts_df["cognate"] = counts_df["antibody"].str.split("_").str[-1].map({"True": True, "False": False})
    
    # Extract bispecific and arm for pairing
    counts_df["bispecific"] = counts_df["antibody"].str.rsplit("_", n=2).str[0]
    counts_df["arm"] = counts_df["antibody"].str.rsplit("_", n=2).str[1]
    
    cognate_counts = counts_df[counts_df["cognate"] == True].sort_values(["bispecific", "arm"]).reset_index(drop=True)
    noncognate_counts = counts_df[counts_df["cognate"] == False].sort_values(["bispecific", "arm"]).reset_index(drop=True)
    
    # Wilcoxon signed-rank test on paired structures
    results = []
    for col in interaction_cols:
        stat, p_value = stats.wilcoxon(cognate_counts[col], noncognate_counts[col], alternative="two-sided")
        results.append({
            "interaction_type": col,
            "cognate_mean": cognate_counts[col].mean(),
            "noncognate_mean": noncognate_counts[col].mean(),
            "difference": cognate_counts[col].mean() - noncognate_counts[col].mean(),
            "p_value": p_value
        })
    
    results_df = pd.DataFrame(results)
    _, results_df["p_adjusted"], _, _ = multipletests(results_df["p_value"], method="fdr_bh")
    results_df = results_df.sort_values("p_adjusted")
    return results_df, counts_df

def plot_interaction_counts(results_df):
    
    plot_df = results_df.reindex(results_df["difference"].abs().sort_values(ascending=False).index)
    
    fig, ax = plt.subplots(figsize=(10, 10))
    
    colors = ["#1A5276" if x > 0 else "#922B21" for x in plot_df["difference"]]
    bars = ax.barh(plot_df["interaction_type"], plot_df["difference"], color=colors)
    
    ax.axvline(x=0, color="black", linestyle="-", linewidth=0.8)
    
    x_range = plot_df["difference"].abs().max()
    min_offset = x_range * 0.03
    
    for i, (_, row) in enumerate(plot_df.iterrows()):
        x_pos = row["difference"]
        
        if row["p_adjusted"] < 0.05:
            p_text = f"p={row['p_adjusted']:.3f}*"
            weight = "bold"
        else:
            p_text = f"p={row['p_adjusted']:.3f}"
            weight = "normal"
        
        if x_pos >= 0:
            ax.text(x_pos + min_offset, i, p_text,
                   va='center', ha='left', fontsize=8, fontweight=weight)
        else:
            ax.text(x_pos - min_offset, i, p_text,
                   va='center', ha='right', fontsize=8, fontweight=weight)
    
    ax.set_xlabel("Mean count difference (Cognate - Non-cognate)", fontsize=11)
    ax.set_title("Interaction type counts: Cognate vs Non-cognate\n"
                "* significant after BH correction (p_adj < 0.05)", fontsize=11)
    
    legend_elements = [Patch(facecolor="#1A5276", label="More frequent in cognate"),
                      Patch(facecolor="#922B21", label="More frequent in non-cognate")]
    ax.legend(handles=legend_elements, fontsize=9)
    
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    
    plt.tight_layout()
    plt.savefig("interaction_counts.png", dpi=150, bbox_inches='tight')
    plt.close()
    print("Saved interaction_counts.png")


## PCA and UMAP outlier investigations -- outdated, need to be re-done ##
def investigate_outliers(feature_matrix, counts_df, embedding, eps=0.5, min_samples=5, method="umap"):
    # Create DataFrame from embedding
    col1, col2 = (f"{method.upper()}1", f"{method.upper()}2")
    embed_df = pd.DataFrame(embedding[:, :2], 
                            columns=[col1, col2], 
                            index=feature_matrix.index)
    
    # Run DBSCAN
    dbscan = DBSCAN(eps=eps, min_samples=min_samples)
    embed_df["cluster"] = dbscan.fit_predict(embedding[:, :2])
    
    print(f"Clusters found: {embed_df['cluster'].unique()}")
    print(f"Cluster sizes:\n{embed_df['cluster'].value_counts()}")
    
    # Identify outlier structures (cluster == -1 means noise in DBSCAN)
    # or structures in minority clusters
    main_cluster = embed_df["cluster"].value_counts().index[0]
    outliers = embed_df[embed_df["cluster"] != main_cluster].index
    print(f"\nNumber of outlier structures: {len(outliers)}")
    print(f"Outlier structures:\n{outliers.tolist()}")
    cognate_status = pd.Series(outliers).str.split("_").str[-1]
    print(f"\nCognate status of outliers:\n{cognate_status.value_counts()}")   

    # Get interaction counts for outliers
    outlier_counts = counts_df[counts_df["antibody"].isin(outliers)]
    cognate_counts = counts_df[counts_df["cognate"] == True]
    noncognate_counts = counts_df[counts_df["cognate"] == False]
    
    interaction_cols = ['AMIDEAMIDE', 'AMIDERING', 'CARBONPI', 'CATIONPI', 'DONORPI', 
                       'EE', 'EF', 'ET', 'FE', 'FF', 'FT', 'METSULPHURPI', 'OE', 'OF', 
                       'OT', 'aromatic', 'carbonyl', 'covalent', 'hbond', 'hydrophobic', 
                       'ionic', 'polar', 'proximal', 'vdw', 'vdw_clash', 'weak_hbond', 
                       'weak_polar']
    
    # Compare means
    comparison = pd.DataFrame({
        "outlier_mean": outlier_counts[interaction_cols].mean(),
        "cognate_mean": cognate_counts[interaction_cols].mean(),
        "noncognate_mean": noncognate_counts[interaction_cols].mean(),
    })
    comparison["outlier_vs_cognate_diff"] = comparison["outlier_mean"] - comparison["cognate_mean"]
    comparison["outlier_vs_noncognate_diff"] = comparison["outlier_mean"] - comparison["noncognate_mean"]
    
    print("\nInteraction count comparison:")
    print(comparison.sort_values("outlier_vs_cognate_diff", ascending=False).to_string())
    
    comparison.to_csv(f"outlier_interaction_counts_{method}.csv")
    
    # Plot embedding with clusters highlighted
    plt.figure(figsize=(10, 8))
    scatter = plt.scatter(embedding[:, 0], embedding[:, 1], 
                         c=embed_df["cluster"], cmap="tab10", alpha=0.5)
    plt.colorbar(scatter, label="Cluster")
    plt.xlabel(col1)
    plt.ylabel(col2)
    plt.title(f"{method.upper()} coloured by DBSCAN cluster")
    plt.savefig(f"{method}_dbscan_clusters.png")
    plt.close()
    
    return comparison, outliers, embed_df


## Get IMGT region for plots ##
def get_imgt_region(position_str):
    """Assign IMGT structural region to a position"""
    try:
        # Extract numeric part
        pos = int(''.join(filter(lambda x: x.isdigit(), str(position_str))))
    except (ValueError, TypeError):
        return "Unknown"
    
    # IMGT region definitions for VH and VL
    if 1 <= pos <= 26:
        return "FWR1"
    elif 27 <= pos <= 38:
        return "CDR1"
    elif 39 <= pos <= 55:
        return "FWR2"
    elif 56 <= pos <= 65:
        return "CDR2"
    elif 66 <= pos <= 104:
        return "FWR3"
    elif 105 <= pos <= 117:
        return "CDR3"
    elif 118 <= pos <= 128:
        return "FWR4"
    else:
        return "Unknown"
    
## Calculate per IMGT position - the frequency of different interaction types ##
def calculate_position_interaction_table(bispec_interactions, chain_id, cognate=True):
    
    interaction_cols = ['hbond', 'hydrophobic', 'ionic', 'polar', 'vdw',
                       'vdw_clash', 'aromatic', 'weak_hbond', 'weak_polar',
                       'proximal', 'carbonyl', 'AMIDEAMIDE', 'AMIDERING',
                       'CARBONPI', 'CATIONPI', 'DONORPI', 'EE', 'EF', 'ET',
                       'FE', 'FF', 'FT', 'METSULPHURPI', 'OE', 'OF', 'OT']
    interaction_cols = [c for c in interaction_cols if c in bispec_interactions.columns]
    
    df = bispec_interactions.copy()
    df["bgn_resnum"] = df["bgn_resnum"].astype(str) + df["bgn_ins_code"].str.strip().fillna("")
    df["end_resnum"] = df["end_resnum"].astype(str) + df["end_ins_code"].str.strip().fillna("")
    df["cognate_flag"] = df["antibody"].str.split("_").str[-1].map({"True": True, "False": False})
    df = df[df["cognate_flag"] == cognate]
    
    n_structures = df["antibody"].nunique()
    
    bgn_rows = df[df["bgn_chain"] == chain_id].copy()
    bgn_rows["imgt_pos"] = bgn_rows["bgn_resnum"]
    
    end_rows = df[df["end_chain"] == chain_id].copy()
    end_rows["imgt_pos"] = end_rows["end_resnum"]
    
    combined = pd.concat([bgn_rows, end_rows])
    
    # Step 1: Get raw counts per position
    position_counts = combined.groupby("imgt_pos")[interaction_cols].sum()
    
    # Step 2: Convert to percentages of total interactions at that position
    interaction_sum = position_counts.sum(axis=1)
    for col in interaction_cols:
        position_counts[col] = (position_counts[col] / interaction_sum * 100).round(1)
    position_counts[interaction_cols] = position_counts[interaction_cols].fillna(0)
    
    # Step 3: Add total_interactions as proportion of structures with contact
    contact_freq = combined.groupby("imgt_pos")["antibody"].nunique() / n_structures
    position_counts.insert(0, "total_interactions", contact_freq)
    
    position_counts = position_counts.round(3)
    
    # Step 4: Include all IMGT positions 1-128 plus insertion codes
    all_positions = [str(i) for i in range(1, 129)]
    observed_positions = set(position_counts.index)
    insertion_positions = [p for p in observed_positions if not p.isdigit()]
    all_positions = sorted(
        set(all_positions) | set(insertion_positions),
        key=lambda x: (int(''.join(filter(str.isdigit, x))), 
                       ''.join(filter(str.isalpha, x)))
    )
    
    # Step 5: Reindex to include all positions, fill missing with 0
    position_counts = position_counts.reindex(all_positions, fill_value=0)
    position_counts.index.name = "position"
    
    # Step 6: Filter to only positions with contacts
    position_counts = position_counts[position_counts["total_interactions"] > 0]
    
    # Step 7: Add IMGT region column as leftmost column
    position_counts.insert(0, "imgt_structural_region",
                          position_counts.index.map(get_imgt_region))
    
    return position_counts

## Saves as colour coded excel ##
def style_and_save_table(table, chain_id, label):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    
    numeric_cols = [c for c in table.columns if c != "imgt_structural_region"]
    table_numeric = table[numeric_cols]
    table_numeric = table_numeric.loc[:, (table_numeric > 0).any()]
    table_plot = pd.concat([table[["imgt_structural_region"]], table_numeric], axis=1)
    
    csv_path = f"V{chain_id}_position_interactions_{label}.csv"
    table_plot.to_csv(csv_path)
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"V{chain_id}_{label}"
    
    headers = ["position"] + list(table_plot.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = PatternFill("solid", start_color="2F4F4F", fgColor="2F4F4F")
        cell.font = Font(bold=True, name='Arial', size=9, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    for row_idx, (pos, row) in enumerate(table_plot.iterrows(), 2):
        region = row["imgt_structural_region"]
        
        if "CDR" in str(region):
            row_bg = "FFE5E5"
        elif "FWR" in str(region):
            row_bg = "E5F0FF"
        else:
            row_bg = "FFFFFF"
        
        # Position column
        cell = ws.cell(row=row_idx, column=1, value=str(pos))
        cell.fill = PatternFill("solid", start_color=row_bg, fgColor=row_bg)
        cell.font = Font(name='Arial', size=9)
        cell.alignment = Alignment(horizontal='center')
        
        for col_idx, col_name in enumerate(table_plot.columns, 2):
            value = row[col_name]
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name='Arial', size=9)
            cell.alignment = Alignment(horizontal='center')
            
            if col_name == "imgt_structural_region":
                # Row background colour only
                cell.fill = PatternFill("solid", start_color=row_bg, fgColor=row_bg)
                cell.font = Font(name='Arial', size=9, bold=True)
            
            elif col_name == "total_interactions":
                # Green gradient: white (0) to dark green (1.0)
                intensity = float(value) if value else 0
                r = int(255 - intensity * 150)
                g = 255
                b = int(255 - intensity * 150)
                hex_color = f"{r:02X}{g:02X}{b:02X}"
                cell.fill = PatternFill("solid", start_color=hex_color, fgColor=hex_color)
                cell.value = round(float(value), 3) if value else 0
            
            else:
                # Grey gradient: white (0%) to dark grey (100%)
                intensity = float(value) / 100 if value else 0
                grey = int(255 - intensity * 180)
                hex_color = f"{grey:02X}{grey:02X}{grey:02X}"
                cell.fill = PatternFill("solid", start_color=hex_color, fgColor=hex_color)
                # Use white font for dark cells so text is readable
                font_color = "FFFFFF" if intensity > 0.6 else "000000"
                cell.font = Font(name='Arial', size=9, color=font_color)
                cell.value = round(float(value), 1) if value else 0
    
    ws.column_dimensions[get_column_letter(1)].width = 10
    ws.column_dimensions[get_column_letter(2)].width = 20
    ws.column_dimensions[get_column_letter(3)].width = 18
    for col_idx in range(4, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12
    
    ws.freeze_panes = "A2"
    
    xlsx_path = f"V{chain_id}_position_interactions_{label}.xlsx"
    
    try:
        wb.save(xlsx_path)
        print(f"Saved {xlsx_path}")
    except Exception as e:
        print(f"Error saving xlsx: {e}")
    
    print(f"\nInteraction counts per IMGT position in {label} structures (V{chain_id}):")
    print(table_plot[table_plot["total_interactions"] > 0].to_string())
    print(f"Saved {csv_path}")
    
    return table_plot

## Count how often each position pair appears across all structures BUT doesn't separate by cognate status ##
def calculate_all_contact_frequencies(bispec_interactions):
    """Calculate contact frequency for all IMGT position pairs across all structures
    to replicate Dudzic et al Figure 4/5"""
    
    df = bispec_interactions.copy()
    df["bgn_resnum"] = df["bgn_resnum"].astype(str) + df["bgn_ins_code"].str.strip().fillna("")
    df["end_resnum"] = df["end_resnum"].astype(str) + df["end_ins_code"].str.strip().fillna("")
    df["res_pair"] = df["bgn_chain"] + df["bgn_resnum"] + "_" + df["end_chain"] + df["end_resnum"]
    
    n_structures = df["antibody"].nunique()
    
    # Count how often each position pair appears across all structures
    contact_freq = df.groupby("res_pair")["antibody"].nunique() / n_structures
    contact_freq = contact_freq.sort_values(ascending=False)
    
    # Also get individual position frequencies (like Dudzic Fig 4)
    bgn_freq = df.groupby(["bgn_chain", "bgn_resnum"])["antibody"].nunique() / n_structures
    end_freq = df.groupby(["end_chain", "end_resnum"])["antibody"].nunique() / n_structures
    
    return contact_freq, bgn_freq, end_freq

## Probably don't need this anymore but plots the above ##
def plot_contact_frequencies(contact_freq, bgn_freq, end_freq, top_n=30):
    """Plot contact frequencies similar to Dudzic et al"""
    '''Probably don't need this anymore'''
    
    # Plot top position pairs
    plt.figure(figsize=(12, 8))
    top_pairs = contact_freq.head(top_n)
    plt.barh(range(len(top_pairs)), top_pairs.values)
    plt.yticks(range(len(top_pairs)), top_pairs.index, fontsize=8)
    plt.xlabel("Frequency (proportion of structures)")
    plt.title(f"Top {top_n} most frequent VH-VL contact position pairs")
    plt.tight_layout()
    plt.savefig("contact_pair_frequencies.png")
    plt.close()
    
    # Plot individual VH position frequencies
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 6))
    
    vh_freq = bgn_freq[bgn_freq.index.get_level_values("bgn_chain") == "H"].sort_values(ascending=False).head(top_n)
    vl_freq = bgn_freq[bgn_freq.index.get_level_values("bgn_chain") == "L"].sort_values(ascending=False).head(top_n)
    
    ax1.bar(range(len(vh_freq)), vh_freq.values)
    ax1.set_xticks(range(len(vh_freq)))
    ax1.set_xticklabels([x[1] for x in vh_freq.index], rotation=90, fontsize=8)
    ax1.set_title("VH position contact frequencies")
    ax1.set_ylabel("Frequency")
    
    ax2.bar(range(len(vl_freq)), vl_freq.values)
    ax2.set_xticks(range(len(vl_freq)))
    ax2.set_xticklabels([x[1] for x in vl_freq.index], rotation=90, fontsize=8)
    ax2.set_title("VL position contact frequencies")
    ax2.set_ylabel("Frequency")
    
    plt.tight_layout()
    plt.savefig("individual_position_frequencies.png")
    plt.close()


## Per IMGT ID contact frequency - separate for VH, VL, cognate, noncognate ##
def plot_imgt_contact_frequency(bispec_interactions, chain_id, cognate, label, sig_positions=None):
    """
    Plot contact frequency per IMGT position like Dudzic Fig 4.
    Y-axis: proportion of structures where that position is in contact
    X-axis: all IMGT positions 1-128 including insertion codes
    Coloured by IMGT region
    """
    
    # IMGT region colours
    region_colors = {
        "FWR1": "#2196F3",  # blue
        "CDR1": "#F44336",  # red
        "FWR2": "#2196F3",  # green
        "CDR2": "#F44336",  # orange
        "FWR3": "#2196F3",  # purple
        "CDR3": "#F44336",  # pink
        "FWR4": "#2196F3",  # cyan
        "Unknown": "#9E9E9E"  # grey
    }
    
    df = bispec_interactions.copy()
    df["bgn_resnum"] = df["bgn_resnum"].astype(str) + df["bgn_ins_code"].str.strip().fillna("")
    df["end_resnum"] = df["end_resnum"].astype(str) + df["end_ins_code"].str.strip().fillna("")
    df["cognate_flag"] = df["antibody"].str.split("_").str[-1].map({"True": True, "False": False})
    df = df[df["cognate_flag"] == cognate]
    
    n_structures = df["antibody"].nunique()
    
    # Get rows where chain_id is bgn or end
    bgn_rows = df[df["bgn_chain"] == chain_id].copy()
    bgn_rows["imgt_pos"] = bgn_rows["bgn_resnum"]
    
    end_rows = df[df["end_chain"] == chain_id].copy()
    end_rows["imgt_pos"] = end_rows["end_resnum"]
    
    combined = pd.concat([bgn_rows, end_rows])
    
    # For each position, count number of structures where it is in contact
    pos_structures = combined.groupby("imgt_pos")["antibody"].nunique()
    
    # Normalise by total number of structures
    contact_freq = pos_structures / n_structures
    
    # Generate all IMGT positions 1-128 plus common insertion codes
    all_positions = []
    for i in range(1, 129):
        all_positions.append(str(i))
    # Add common insertion codes for CDR3 region
    for suffix in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        for base in [111, 112]:
            all_positions.append(f"{base}{suffix}")
    
    # Sort positions numerically with insertion codes after their base number
    def sort_key(pos):
        digits = ''.join(filter(str.isdigit, pos))
        letters = ''.join(filter(str.isalpha, pos))
        return (int(digits) if digits else 0, letters)
    
    all_positions = sorted(set(all_positions) | set(contact_freq.index), key=sort_key)
    
    # Build frequency series for all positions
    freq_series = pd.Series(0.0, index=all_positions)
    for pos in contact_freq.index:
        if pos in freq_series.index:
            freq_series[pos] = contact_freq[pos]
    
    # Assign regions and colours
    colors = [region_colors[get_imgt_region(pos)] for pos in freq_series.index]
    
    # Plot
    fig, ax = plt.subplots(figsize=(24, 6))

    bars = ax.bar(range(len(freq_series)), freq_series.values,
                  color=colors, width=0.8, edgecolor='none')
    
    # Add stars above significant positions
    if sig_positions:
        for i, pos in enumerate(freq_series.index):
            if pos in sig_positions:
                ax.text(i, freq_series[pos] + 0.02, '*', 
                       ha='center', va='bottom', fontsize=10, 
                       color='black', fontweight='bold')
        
    ax.set_xticks(range(len(freq_series)))
    ax.set_xticklabels(freq_series.index, rotation=90, fontsize=5)
    ax.set_ylabel("Contact Frequency\n(proportion of structures)", fontsize=11)
    ax.set_xlabel("IMGT Position", fontsize=11)
    ax.set_title(f"V{chain_id} IMGT Position Contact Frequencies ({label})\nn={n_structures} structures", 
                fontsize=13)
    ax.set_ylim(0, 1.05)
    
    # Add region boundary lines and labels
    current_region = None
    region_start = 0
    region_labels = []
    
    for i, pos in enumerate(freq_series.index):
        region = get_imgt_region(pos)
        if region != current_region:
            if current_region is not None:
                region_labels.append((current_region, region_start, i - 1))
                ax.axvline(x=i - 0.5, color='black', linewidth=0.5, linestyle='--', alpha=0.3)
            current_region = region
            region_start = i
    region_labels.append((current_region, region_start, len(freq_series) - 1))
    
    # Add region labels below x-axis
    for region, start, end in region_labels:
        mid = (start + end) / 2
        ax.text(mid, -0.12, region, ha='center', va='top', 
               fontsize=7, transform=ax.get_xaxis_transform(),
               color=region_colors.get(region, 'black'), fontweight='bold')
    
    # Legend
    legend_handles = [Patch(color=color, label=region) 
                     for region, color in region_colors.items() 
                     if region != "Unknown"]
    ax.legend(handles=legend_handles, loc='upper right', fontsize=9, 
             title="IMGT Region", title_fontsize=10)
    
    plt.tight_layout()
    outpath = f"V{chain_id}_contact_frequency_{label}.png"
    plt.savefig(outpath, dpi=150, bbox_inches='tight')
    plt.close()
    print(f"Saved {outpath}")


## Stats on difference in total contact counts per IMGT pos in cognate vs noncognate ##
def test_imgt_position_differences(bispec_interactions, chain_id):
    from statsmodels.stats.multitest import multipletests

    df = bispec_interactions.copy()
    df["bgn_resnum"] = df["bgn_resnum"].astype(str) + df["bgn_ins_code"].str.strip().fillna("")
    df["end_resnum"] = df["end_resnum"].astype(str) + df["end_ins_code"].str.strip().fillna("")
    df["cognate_flag"] = df["antibody"].str.split("_").str[-1].map({"True": True, "False": False})
    df["bispecific"] = df["antibody"].str.rsplit("_", n=2).str[0]
    df["arm"] = df["antibody"].str.rsplit("_", n=2).str[1]

    bgn_rows = df[df["bgn_chain"] == chain_id].copy()
    bgn_rows["imgt_pos"] = bgn_rows["bgn_resnum"]
    end_rows = df[df["end_chain"] == chain_id].copy()
    end_rows["imgt_pos"] = end_rows["end_resnum"]
    combined = pd.concat([bgn_rows, end_rows])

    # Build per-structure binary matrix
    contact_binary = combined.groupby(["antibody", "imgt_pos"])["imgt_pos"].count().unstack(fill_value=0)
    contact_binary = (contact_binary > 0).astype(int)

    contact_binary["cognate_flag"] = contact_binary.index.map(
        lambda x: x.split("_")[-1] == "True"
    )
    contact_binary["bispecific"] = contact_binary.index.map(
        lambda x: "_".join(x.rsplit("_", 2)[:-2])
    )
    contact_binary["arm"] = contact_binary.index.map(
        lambda x: x.rsplit("_", 2)[-2]
    )

    cognate_df = contact_binary[contact_binary["cognate_flag"] == True]\
        .sort_values(["bispecific", "arm"]).reset_index(drop=True)
    noncognate_df = contact_binary[contact_binary["cognate_flag"] == False]\
        .sort_values(["bispecific", "arm"]).reset_index(drop=True)

    imgt_positions = [c for c in contact_binary.columns
                     if c not in ["cognate_flag", "bispecific", "arm"]]

    results = []
    p_values = []

    for pos in imgt_positions:
        cog_vals = cognate_df[pos].values
        noncog_vals = noncognate_df[pos].values

        # McNemar's contingency table
        b = ((cog_vals == 1) & (noncog_vals == 0)).sum()
        c = ((cog_vals == 0) & (noncog_vals == 1)).sum()

        if b + c == 0:
            p_value = 1.0
        else:
            if b + c < 25:
                result = stats.binomtest(b, b + c, p=0.5, alternative="two-sided")
                p_value = result.pvalue
            else:
                chi2 = (abs(b - c) - 1) ** 2 / (b + c)
                p_value = stats.chi2.sf(chi2, df=1)

        p_values.append(p_value)
        results.append({
            "imgt_pos": pos,
            "cognate_freq": cog_vals.mean(),
            "noncognate_freq": noncog_vals.mean(),
            "n_cog1_noncog0": b,
            "n_cog0_noncog1": c,
            "p_value": p_value
        })

    results_df = pd.DataFrame(results)

    # BH correction across all positions
    _, results_df["p_adjusted"], _, _ = multipletests(p_values, method="fdr_bh")

    n_sig = (results_df["p_adjusted"] < 0.05).sum()
    print(f"V{chain_id}: {n_sig} significant positions after BH correction (McNemar's test)")

    sig_positions = set(results_df[results_df["p_adjusted"] < 0.05]["imgt_pos"].astype(str))
    results_df.to_csv(f"V{chain_id}_imgt_position_stats.csv", index=False)

    return sig_positions, results_df

## Stats on difference in interaction types per region ##
def test_interaction_types_per_region(bispec_interactions, chain_id):
    """
    For each IMGT region, test whether specific interaction types differ
    between cognate and non-cognate structures using Wilcoxon signed-rank test.
    """
    from statsmodels.stats.multitest import multipletests

    interaction_cols = ['hbond', 'hydrophobic', 'ionic', 'polar', 'vdw',
                       'vdw_clash', 'aromatic', 'weak_hbond', 'weak_polar',
                       'proximal', 'carbonyl', 'AMIDEAMIDE', 'AMIDERING',
                       'CARBONPI', 'CATIONPI', 'DONORPI', 'EE', 'EF', 'ET',
                       'FE', 'FF', 'FT', 'METSULPHURPI', 'OE', 'OF', 'OT']
    interaction_cols = [c for c in interaction_cols if c in bispec_interactions.columns]
    regions = ["FWR1", "CDR1", "FWR2", "CDR2", "FWR3", "CDR3", "FWR4"]

    df = bispec_interactions.copy()
    df["bgn_resnum"] = df["bgn_resnum"].astype(str) + df["bgn_ins_code"].str.strip().fillna("")
    df["end_resnum"] = df["end_resnum"].astype(str) + df["end_ins_code"].str.strip().fillna("")
    df["cognate_flag"] = df["antibody"].str.split("_").str[-1].map({"True": True, "False": False})
    df["bispecific"] = df["antibody"].str.rsplit("_", n=2).str[0]
    df["arm"] = df["antibody"].str.rsplit("_", n=2).str[1]

    # Get rows where chain_id is bgn or end and assign imgt_pos and region
    bgn_rows = df[df["bgn_chain"] == chain_id].copy()
    bgn_rows["imgt_pos"] = bgn_rows["bgn_resnum"]
    end_rows = df[df["end_chain"] == chain_id].copy()
    end_rows["imgt_pos"] = end_rows["end_resnum"]
    combined = pd.concat([bgn_rows, end_rows])
    combined["region"] = combined["imgt_pos"].map(get_imgt_region)

    # Sum interaction counts per structure per region
    region_counts_raw = combined.groupby(
        ["antibody", "cognate_flag", "bispecific", "arm", "region"]
    )[interaction_cols].sum().reset_index()

    # Ensure all structures appear for every region by creating a full index
    all_structures = df[["antibody", "cognate_flag", "bispecific", "arm"]].drop_duplicates()
    full_index = pd.MultiIndex.from_product(
        [all_structures["antibody"].values, regions],
        names=["antibody", "region"]
    )
    full_df = pd.DataFrame(index=full_index).reset_index()
    full_df = full_df.merge(all_structures, on="antibody", how="left")

    region_counts = full_df.merge(
        region_counts_raw,
        on=["antibody", "cognate_flag", "bispecific", "arm", "region"],
        how="left"
    ).fillna(0)

    all_results = []

    for region in regions:
        region_df = region_counts[region_counts["region"] == region]

        cognate_df = region_df[region_df["cognate_flag"] == True]\
            .sort_values(["bispecific", "arm"]).reset_index(drop=True)
        noncognate_df = region_df[region_df["cognate_flag"] == False]\
            .sort_values(["bispecific", "arm"]).reset_index(drop=True)

        if len(cognate_df) == 0 or len(noncognate_df) == 0:
            continue

        p_values = []
        region_results = []

        for col in interaction_cols:
            cog_vals = cognate_df[col].values
            noncog_vals = noncognate_df[col].values

            if (cog_vals == noncog_vals).all():
                p_values.append(1.0)
                region_results.append({
                    "region": region,
                    "interaction_type": col,
                    "cognate_mean": round(cog_vals.mean(), 3),
                    "noncognate_mean": round(noncog_vals.mean(), 3),
                    "difference": round(cog_vals.mean() - noncog_vals.mean(), 3),
                    "p_value": 1.0
                })
            else:
                try:
                    _, p = stats.wilcoxon(cog_vals, noncog_vals, alternative="two-sided")
                    p_values.append(p)
                    region_results.append({
                        "region": region,
                        "interaction_type": col,
                        "cognate_mean": round(cog_vals.mean(), 3),
                        "noncognate_mean": round(noncog_vals.mean(), 3),
                        "difference": round(cog_vals.mean() - noncog_vals.mean(), 3),
                        "p_value": round(p, 4)
                    })
                except ValueError:
                    p_values.append(1.0)
                    region_results.append({
                        "region": region,
                        "interaction_type": col,
                        "cognate_mean": round(cog_vals.mean(), 3),
                        "noncognate_mean": round(noncog_vals.mean(), 3),
                        "difference": round(cog_vals.mean() - noncog_vals.mean(), 3),
                        "p_value": 1.0
                    })

        # BH correction within each region
        _, p_adjusted, _, _ = multipletests(p_values, method="fdr_bh")
        for i, result in enumerate(region_results):
            result["p_adjusted"] = round(p_adjusted[i], 4)
            result["significant"] = p_adjusted[i] < 0.05

        all_results.extend(region_results)

    results_df = pd.DataFrame(all_results)
    results_df.to_csv(f"V{chain_id}_interaction_types_per_region.csv", index=False)

    n_sig = results_df["significant"].sum()
    print(f"\nV{chain_id} interaction types per region:")
    print(f"Significant results: {n_sig}")
    if n_sig > 0:
        print(results_df[results_df["significant"]].to_string())

    return results_df

## Slope chart - don't need this##
def plot_slope_chart_v2(bispec_interactions, output_prefix="slope_chart"):
    
    def get_contact_pairs(df, cognate):
        d = df.copy()
        d["bgn_resnum"] = d["bgn_resnum"].astype(str) + d["bgn_ins_code"].str.strip().fillna("")
        d["end_resnum"] = d["end_resnum"].astype(str) + d["end_ins_code"].str.strip().fillna("")
        d["cognate_flag"] = d["antibody"].str.split("_").str[-1].map({"True": True, "False": False})
        d = d[d["cognate_flag"] == cognate]
        n_structures = d["antibody"].nunique()
        
        bgn_h = d[d["bgn_chain"] == "H"].copy()
        bgn_h["vh_pos"] = bgn_h["bgn_resnum"]
        bgn_h["vl_pos"] = bgn_h["end_resnum"]
        
        end_h = d[d["end_chain"] == "H"].copy()
        end_h["vh_pos"] = end_h["end_resnum"]
        end_h["vl_pos"] = end_h["bgn_resnum"]
        
        combined = pd.concat([bgn_h, end_h])
        pair_freq = combined.groupby(["vh_pos", "vl_pos"])["antibody"].nunique() / n_structures
        
        return pair_freq, n_structures
    
    def sort_key(pos):
        digits = ''.join(filter(str.isdigit, str(pos)))
        letters = ''.join(filter(str.isalpha, str(pos)))
        return (int(digits) if digits else 0, letters)
    
    def get_all_positions(df, chain_id):
        """Get all positions present in the data for a chain, 
        plus all standard IMGT positions 1-128"""
        d = df.copy()
        d["bgn_resnum"] = d["bgn_resnum"].astype(str) + d["bgn_ins_code"].str.strip().fillna("")
        d["end_resnum"] = d["end_resnum"].astype(str) + d["end_ins_code"].str.strip().fillna("")
        
        bgn_pos = set(d[d["bgn_chain"] == chain_id]["bgn_resnum"].unique())
        end_pos = set(d[d["end_chain"] == chain_id]["end_resnum"].unique())
        observed = bgn_pos | end_pos
        
        # Add all standard IMGT positions 1-128
        standard = {str(i) for i in range(1, 129)}
        all_pos = observed | standard
        
        return sorted(all_pos, key=sort_key)
    
    def get_region_color(pos):
        region = get_imgt_region(pos)
        if "CDR" in region:
            return (0.96, 0.26, 0.21, 0.3)
        else:
            return (0.13, 0.59, 0.95, 0.3)
    
    # Get contact pairs
    cog_pairs, n_cog = get_contact_pairs(bispec_interactions, True)
    noncog_pairs, n_noncog = get_contact_pairs(bispec_interactions, False)
    
    # Find positions that differ
    threshold = 0.1
    cog_pair_set = set(cog_pairs[cog_pairs >= threshold].index)
    noncog_pair_set = set(noncog_pairs[noncog_pairs >= threshold].index)
    
    different_vh = set()
    different_vl = set()
    for pair in cog_pair_set.symmetric_difference(noncog_pair_set):
        different_vh.add(pair[0])
        different_vl.add(pair[1])
    for pair in cog_pair_set | noncog_pair_set:
        cog_freq = cog_pairs.get(pair, 0)
        noncog_freq = noncog_pairs.get(pair, 0)
        if abs(cog_freq - noncog_freq) > threshold:
            different_vh.add(pair[0])
            different_vl.add(pair[1])
    
    # Get ALL positions including 1-128
    all_vh_positions = get_all_positions(bispec_interactions, "H")
    all_vl_positions = get_all_positions(bispec_interactions, "L")
    
    for cognate, label, pair_freq in [
        (True, "cognate", cog_pairs),
        (False, "noncognate", noncog_pairs)
    ]:
        pair_freq_filtered = pair_freq[pair_freq >= threshold]
        
        vh_positions = all_vh_positions
        vl_positions = all_vl_positions
        
        vh_y = {pos: i for i, pos in enumerate(vh_positions)}
        vl_y = {pos: i for i, pos in enumerate(vl_positions)}
        
        n_vh = len(vh_positions)
        n_vl = len(vl_positions)
        max_n = max(n_vh, n_vl)
        
        fig, ax = plt.subplots(figsize=(10, max(20, max_n * 0.25)))
        
        # Draw VH background stripes and labels
        for pos in vh_positions:
            y = vh_y[pos] / max(n_vh - 1, 1)
            region_color = get_region_color(pos)
            stripe_height = 1.0 / max(n_vh - 1, 1)
            
            ax.barh(y, 0.08, left=0.0, 
                   color=region_color[:3],
                   alpha=region_color[3] * 3, 
                   height=stripe_height,
                   align='center')
            
            fontweight = 'bold' if pos in different_vh else 'normal'
            fontsize = 5 if max_n > 100 else 7
            ax.text(-0.02, y, str(pos), ha='right', va='center',
                   fontsize=fontsize, fontweight=fontweight)
        
        # Draw VL background stripes and labels
        for pos in vl_positions:
            y = vl_y[pos] / max(n_vl - 1, 1)
            region_color = get_region_color(pos)
            stripe_height = 1.0 / max(n_vl - 1, 1)
            
            ax.barh(y, 0.08, left=0.92,
                   color=region_color[:3],
                   alpha=region_color[3] * 3,
                   height=stripe_height,
                   align='center')
            
            fontweight = 'bold' if pos in different_vl else 'normal'
            fontsize = 5 if max_n > 100 else 7
            ax.text(1.02, y, str(pos), ha='left', va='center',
                   fontsize=fontsize, fontweight=fontweight)
        
        # Vertical lines
        ax.axvline(x=0.08, color='black', linewidth=1.5)
        ax.axvline(x=0.92, color='black', linewidth=1.5)
        
        # Draw contact lines
        for (vh_pos, vl_pos), freq in pair_freq_filtered.items():
            if vh_pos in vh_y and vl_pos in vl_y:
                y1 = vh_y[vh_pos] / max(n_vh - 1, 1)
                y2 = vl_y[vl_pos] / max(n_vl - 1, 1)
                
                is_cdr = "CDR" in get_imgt_region(vh_pos) or "CDR" in get_imgt_region(vl_pos)
                color = '#E53935' if is_cdr else '#1E88E5'
                alpha = min(0.2 + freq * 0.6, 0.9)
                lw = 0.5 + freq * 2
                
                ax.plot([0.08, 0.92], [y1, y2],
                       color=color, alpha=alpha, linewidth=lw)
        
        # Headers
        ax.text(0.08, 1.01, 'Heavy\nchain', ha='center',
               transform=ax.transAxes, fontsize=11, fontweight='bold')
        ax.text(0.92, 1.01, 'Light\nchain', ha='center',
               transform=ax.transAxes, fontsize=11, fontweight='bold')
        
        # Legend
        fwr_patch = Patch(color='#1E88E5', alpha=0.3, label='Framework')
        cdr_patch = Patch(color='#E53935', alpha=0.3, label='CDR')
        bold_line = mlines.Line2D([], [], color='black', linewidth=2,
                                  label='Bold = differs between\ncognate & non-cognate')
        ax.legend(handles=[fwr_patch, cdr_patch, bold_line],
                 loc='upper center', fontsize=9,
                 bbox_to_anchor=(0.5, 1.0))
        
        ax.set_xlim(-0.15, 1.15)
        ax.set_ylim(-0.02, 1.02)
        ax.set_title(f"VH-VL Inter-chain Contacts ({label})\n"
                    f"Line thickness = contact frequency | "
                    f"Bold = differs cognate vs non-cognate",
                    fontsize=11)
        ax.axis('off')
        
        plt.tight_layout()
        outpath = f"{output_prefix}_{label}.png"
        plt.savefig(outpath, dpi=150, bbox_inches='tight')
        plt.close()
        print(f"Saved {outpath}")


## Condensed slope chart ##
def plot_slope_chart_condensed(bispec_interactions, output_prefix="slope_chart_condensed", interval=10):
    """
    Condensed slope chart showing IMGT positions at specified intervals
    """
    
    def get_contact_pairs(df, cognate):
        d = df.copy()
        d["bgn_resnum"] = d["bgn_resnum"].astype(str) + d["bgn_ins_code"].str.strip().fillna("")
        d["end_resnum"] = d["end_resnum"].astype(str) + d["end_ins_code"].str.strip().fillna("")
        d["cognate_flag"] = d["antibody"].str.split("_").str[-1].map({"True": True, "False": False})
        d = d[d["cognate_flag"] == cognate]
        n_structures = d["antibody"].nunique()
        
        bgn_h = d[d["bgn_chain"] == "H"].copy()
        bgn_h["vh_pos"] = bgn_h["bgn_resnum"]
        bgn_h["vl_pos"] = bgn_h["end_resnum"]
        
        end_h = d[d["end_chain"] == "H"].copy()
        end_h["vh_pos"] = end_h["end_resnum"]
        end_h["vl_pos"] = end_h["bgn_resnum"]
        
        combined = pd.concat([bgn_h, end_h])
        pair_freq = combined.groupby(["vh_pos", "vl_pos"])["antibody"].nunique() / n_structures
        
        return pair_freq, n_structures
    
    def sort_key(pos):
        digits = ''.join(filter(str.isdigit, str(pos)))
        letters = ''.join(filter(str.isalpha, str(pos)))
        return (int(digits) if digits else 0, letters)
    
    def get_region_color(pos):
        region = get_imgt_region(pos)
        if "CDR" in region:
            return (0.96, 0.26, 0.21, 0.3)
        else:
            return (0.13, 0.59, 0.95, 0.3)
    
    # Get contact pairs
    cog_pairs, n_cog = get_contact_pairs(bispec_interactions, True)
    noncog_pairs, n_noncog = get_contact_pairs(bispec_interactions, False)
    
    # Find positions that differ
    threshold = 0.1
    cog_pair_set = set(cog_pairs[cog_pairs >= threshold].index)
    noncog_pair_set = set(noncog_pairs[noncog_pairs >= threshold].index)
    
    different_vh = set()
    different_vl = set()
    for pair in cog_pair_set.symmetric_difference(noncog_pair_set):
        different_vh.add(pair[0])
        different_vl.add(pair[1])
    for pair in cog_pair_set | noncog_pair_set:
        cog_freq = cog_pairs.get(pair, 0)
        noncog_freq = noncog_pairs.get(pair, 0)
        if abs(cog_freq - noncog_freq) > threshold:
            different_vh.add(pair[0])
            different_vl.add(pair[1])
    
    # Use only positions at specified intervals (1, 10, 20, ... 120, 128)
    interval_positions = [str(i) for i in range(1, 129, interval)]
    if "128" not in interval_positions:
        interval_positions.append("128")
    interval_positions = sorted(interval_positions, key=sort_key)
    
    for cognate, label, pair_freq in [
        (True, "cognate", cog_pairs),
        (False, "noncognate", noncog_pairs)
    ]:
        pair_freq_filtered = pair_freq[pair_freq >= threshold]
        
        vh_positions = interval_positions
        vl_positions = interval_positions
        
        vh_y = {pos: i for i, pos in enumerate(vh_positions)}
        vl_y = {pos: i for i, pos in enumerate(vl_positions)}
        
        n_vh = len(vh_positions)
        n_vl = len(vl_positions)
        
        fig, ax = plt.subplots(figsize=(8, 10))
        
        # Draw VH background stripes and labels
        for pos in vh_positions:
            y = vh_y[pos] / max(n_vh - 1, 1)
            region_color = get_region_color(pos)
            stripe_height = 1.0 / max(n_vh - 1, 1)
            
            ax.barh(y, 0.08, left=0.0,
                   color=region_color[:3],
                   alpha=region_color[3] * 3,
                   height=stripe_height,
                   align='center')
            
            fontweight = 'bold' if pos in different_vh else 'normal'
            ax.text(-0.02, y, str(pos), ha='right', va='center',
                   fontsize=9, fontweight=fontweight)
        
        # Draw VL background stripes and labels
        for pos in vl_positions:
            y = vl_y[pos] / max(n_vl - 1, 1)
            region_color = get_region_color(pos)
            stripe_height = 1.0 / max(n_vl - 1, 1)
            
            ax.barh(y, 0.08, left=0.92,
                   color=region_color[:3],
                   alpha=region_color[3] * 3,
                   height=stripe_height,
                   align='center')
            
            fontweight = 'bold' if pos in different_vl else 'normal'
            ax.text(1.02, y, str(pos), ha='left', va='center',
                   fontsize=9, fontweight=fontweight)
        
        # Vertical lines
        ax.axvline(x=0.08, color='black', linewidth=1.5)
        ax.axvline(x=0.92, color='black', linewidth=1.5)
        
        # Draw contact lines - map to nearest interval position
        def nearest_interval(pos):
            try:
                pos_num = int(''.join(filter(str.isdigit, str(pos))))
                interval_nums = [int(p) for p in interval_positions]
                nearest = min(interval_nums, key=lambda x: abs(x - pos_num))
                return str(nearest)
            except (ValueError, TypeError):
                return None
        
        # Aggregate contacts to interval positions
        aggregated = {}
        for (vh_pos, vl_pos), freq in pair_freq_filtered.items():
            vh_interval = nearest_interval(vh_pos)
            vl_interval = nearest_interval(vl_pos)
            if vh_interval and vl_interval:
                key = (vh_interval, vl_interval)
                aggregated[key] = max(aggregated.get(key, 0), freq)
        
        for (vh_pos, vl_pos), freq in aggregated.items():
            if vh_pos in vh_y and vl_pos in vl_y:
                y1 = vh_y[vh_pos] / max(n_vh - 1, 1)
                y2 = vl_y[vl_pos] / max(n_vl - 1, 1)
                
                is_cdr = "CDR" in get_imgt_region(vh_pos) or "CDR" in get_imgt_region(vl_pos)
                color = '#E53935' if is_cdr else '#1E88E5'
                alpha = min(0.2 + freq * 0.6, 0.9)
                lw = 0.5 + freq * 2
                
                ax.plot([0.08, 0.92], [y1, y2],
                       color=color, alpha=alpha, linewidth=lw)
        
        ax.text(0.08, 1.02, 'Heavy\nchain', ha='center',
               transform=ax.transAxes, fontsize=11, fontweight='bold')
        ax.text(0.92, 1.02, 'Light\nchain', ha='center',
               transform=ax.transAxes, fontsize=11, fontweight='bold')
        
        fwr_patch = Patch(color='#1E88E5', alpha=0.3, label='Framework')
        cdr_patch = Patch(color='#E53935', alpha=0.3, label='CDR')
        ax.legend(handles=[fwr_patch, cdr_patch],
                 loc='upper center', fontsize=9,
                 bbox_to_anchor=(0.5, 1.0))
        
        ax.set_xlim(-0.15, 1.15)
        ax.set_ylim(-0.05, 1.05)
        ax.set_title(f"VH-VL Inter-chain Contacts ({label})\n"
                    f"Positions shown at {interval}-residue intervals",
                    fontsize=11)
        ax.axis('off')
        
        plt.tight_layout()
        outpath = f"{output_prefix}_{label}.png"
        plt.savefig(outpath, dpi=150, bbox_inches='tight')
        plt.close()
        print(f"Saved {outpath}")


## Differing pairs between cognate and noncognate slope charts ## 
def get_differing_pairs_table(bispec_interactions, threshold=0.05):
    """
    Generate table of contact pairs that differ between cognate and non-cognate
    """
    
    def get_contact_pairs(df, cognate):
        d = df.copy()
        d["bgn_resnum"] = d["bgn_resnum"].astype(str) + d["bgn_ins_code"].str.strip().fillna("")
        d["end_resnum"] = d["end_resnum"].astype(str) + d["end_ins_code"].str.strip().fillna("")
        d["cognate_flag"] = d["antibody"].str.split("_").str[-1].map({"True": True, "False": False})
        d = d[d["cognate_flag"] == cognate]
        n_structures = d["antibody"].nunique()
        
        bgn_h = d[d["bgn_chain"] == "H"].copy()
        bgn_h["vh_pos"] = bgn_h["bgn_resnum"]
        bgn_h["vl_pos"] = bgn_h["end_resnum"]
        
        end_h = d[d["end_chain"] == "H"].copy()
        end_h["vh_pos"] = end_h["end_resnum"]
        end_h["vl_pos"] = end_h["bgn_resnum"]
        
        combined = pd.concat([bgn_h, end_h])
        pair_freq = combined.groupby(["vh_pos", "vl_pos"])["antibody"].nunique() / n_structures
        
        return pair_freq
    
    cog_pairs = get_contact_pairs(bispec_interactions, True)
    noncog_pairs = get_contact_pairs(bispec_interactions, False)
    
    # Get all pairs present in either group
    all_pairs = set(cog_pairs.index) | set(noncog_pairs.index)
    
    results = []
    for (vh_pos, vl_pos) in all_pairs:
        cog_freq = cog_pairs.get((vh_pos, vl_pos), 0)
        noncog_freq = noncog_pairs.get((vh_pos, vl_pos), 0)
        diff = cog_freq - noncog_freq
        abs_diff = abs(diff)
        
        if abs_diff >= threshold:
            results.append({
                "vh_pos": vh_pos,
                "vl_pos": vl_pos,
                "vh_region": get_imgt_region(vh_pos),
                "vl_region": get_imgt_region(vl_pos),
                "cognate_freq": round(cog_freq, 3),
                "noncognate_freq": round(noncog_freq, 3),
                "difference": round(diff, 3),
                "abs_difference": round(abs_diff, 3),
                "higher_in": "cognate" if diff > 0 else "noncognate"
            })
    
    if len(results) == 0:
        print(f"No pairs differ by more than {threshold} between cognate and non-cognate")
        print("Trying lower threshold of 0.05...")
        # Show top differences regardless of threshold
        all_results = []
        for (vh_pos, vl_pos) in all_pairs:
            cog_freq = cog_pairs.get((vh_pos, vl_pos), 0)
            noncog_freq = noncog_pairs.get((vh_pos, vl_pos), 0)
            diff = cog_freq - noncog_freq
            all_results.append({
                "vh_pos": vh_pos,
                "vl_pos": vl_pos,
                "vh_region": get_imgt_region(vh_pos),
                "vl_region": get_imgt_region(vl_pos),
                "cognate_freq": round(cog_freq, 3),
                "noncognate_freq": round(noncog_freq, 3),
                "difference": round(diff, 3),
                "abs_difference": round(abs(diff), 3),
                "higher_in": "cognate" if diff > 0 else "noncognate"
            })
        results_df = pd.DataFrame(all_results).sort_values("abs_difference", ascending=False)
        print(f"\nTop 20 pairs with largest differences (no threshold):")
        print(results_df.head(20).to_string())
        results_df.to_csv("differing_contact_pairs.csv", index=False)
        return results_df

    results_df = pd.DataFrame(results).sort_values("abs_difference", ascending=False)

## Look at differences in character of region contacts ##
def test_interaction_types_per_region_pair(bispec_interactions):
    """
    For each VH region x VL region pair, test whether specific interaction types
    differ between cognate and non-cognate structures using Wilcoxon signed-rank test.
    """
    from statsmodels.stats.multitest import multipletests

    interaction_cols = ['hbond', 'hydrophobic', 'ionic', 'polar', 'vdw',
                       'vdw_clash', 'aromatic', 'weak_hbond', 'weak_polar',
                       'proximal', 'carbonyl', 'AMIDEAMIDE', 'AMIDERING',
                       'CARBONPI', 'CATIONPI', 'DONORPI', 'EE', 'EF', 'ET',
                       'FE', 'FF', 'FT', 'METSULPHURPI', 'OE', 'OF', 'OT']
    interaction_cols = [c for c in interaction_cols if c in bispec_interactions.columns]
    regions = ["FWR1", "CDR1", "FWR2", "CDR2", "FWR3", "CDR3", "FWR4"]

    df = bispec_interactions.copy()
    df["bgn_resnum"] = df["bgn_resnum"].astype(str) + df["bgn_ins_code"].str.strip().fillna("")
    df["end_resnum"] = df["end_resnum"].astype(str) + df["end_ins_code"].str.strip().fillna("")
    df["cognate_flag"] = df["antibody"].str.split("_").str[-1].map({"True": True, "False": False})
    df["bispecific"] = df["antibody"].str.rsplit("_", n=2).str[0]
    df["arm"] = df["antibody"].str.rsplit("_", n=2).str[1]

    # Standardise so H chain is always bgn
    bgn_h = df[df["bgn_chain"] == "H"].copy()
    bgn_h["vh_pos"] = bgn_h["bgn_resnum"]
    bgn_h["vl_pos"] = bgn_h["end_resnum"]

    end_h = df[df["end_chain"] == "H"].copy()
    end_h["vh_pos"] = end_h["end_resnum"]
    end_h["vl_pos"] = end_h["bgn_resnum"]

    combined = pd.concat([bgn_h, end_h])

    # Assign regions
    combined["vh_region"] = combined["vh_pos"].map(get_imgt_region)
    combined["vl_region"] = combined["vl_pos"].map(get_imgt_region)
    combined["region_pair"] = combined["vh_region"] + "_" + combined["vl_region"]

    # Get all region pairs that actually have contacts
    observed_region_pairs = combined["region_pair"].unique()

    # Sum interaction counts per structure per region pair
    region_pair_counts_raw = combined.groupby(
        ["antibody", "cognate_flag", "bispecific", "arm", "region_pair"]
    )[interaction_cols].sum().reset_index()

    # Ensure all structures appear for every region pair
    all_structures = df[["antibody", "cognate_flag", "bispecific", "arm"]].drop_duplicates()
    full_index = pd.MultiIndex.from_product(
        [all_structures["antibody"].values, observed_region_pairs],
        names=["antibody", "region_pair"]
    )
    full_df = pd.DataFrame(index=full_index).reset_index()
    full_df = full_df.merge(all_structures, on="antibody", how="left")

    region_pair_counts = full_df.merge(
        region_pair_counts_raw,
        on=["antibody", "cognate_flag", "bispecific", "arm", "region_pair"],
        how="left"
    ).fillna(0)

    all_results = []

    for region_pair in sorted(observed_region_pairs):
        rp_df = region_pair_counts[region_pair_counts["region_pair"] == region_pair]

        cognate_df = rp_df[rp_df["cognate_flag"] == True]\
            .sort_values(["bispecific", "arm"]).reset_index(drop=True)
        noncognate_df = rp_df[rp_df["cognate_flag"] == False]\
            .sort_values(["bispecific", "arm"]).reset_index(drop=True)

        if len(cognate_df) == 0 or len(noncognate_df) == 0:
            continue

        p_values = []
        rp_results = []

        for col in interaction_cols:
            cog_vals = cognate_df[col].values
            noncog_vals = noncognate_df[col].values

            if (cog_vals == noncog_vals).all():
                p_values.append(1.0)
                rp_results.append({
                    "region_pair": region_pair,
                    "interaction_type": col,
                    "cognate_mean": round(cog_vals.mean(), 3),
                    "noncognate_mean": round(noncog_vals.mean(), 3),
                    "difference": round(cog_vals.mean() - noncog_vals.mean(), 3),
                    "p_value": 1.0
                })
            else:
                try:
                    _, p = stats.wilcoxon(cog_vals, noncog_vals, alternative="two-sided")
                    p_values.append(p)
                    rp_results.append({
                        "region_pair": region_pair,
                        "interaction_type": col,
                        "cognate_mean": round(cog_vals.mean(), 3),
                        "noncognate_mean": round(noncog_vals.mean(), 3),
                        "difference": round(cog_vals.mean() - noncog_vals.mean(), 3),
                        "p_value": round(p, 4)
                    })
                except ValueError:
                    p_values.append(1.0)
                    rp_results.append({
                        "region_pair": region_pair,
                        "interaction_type": col,
                        "cognate_mean": round(cog_vals.mean(), 3),
                        "noncognate_mean": round(noncog_vals.mean(), 3),
                        "difference": round(cog_vals.mean() - noncog_vals.mean(), 3),
                        "p_value": 1.0
                    })

        # BH correction within each region pair
        _, p_adjusted, _, _ = multipletests(p_values, method="fdr_bh")
        for i, result in enumerate(rp_results):
            result["p_adjusted"] = round(p_adjusted[i], 4)
            result["significant"] = p_adjusted[i] < 0.05

        all_results.extend(rp_results)

    results_df = pd.DataFrame(all_results)
    results_df.to_csv("region_pair_interaction_types.csv", index=False)

    n_sig = results_df["significant"].sum()
    print(f"\nRegion pair interaction types:")
    print(f"Significant results: {n_sig}")
    if n_sig > 0:
        print(results_df[results_df["significant"]].to_string())
    else:
        print("No significant results after BH correction")
        print("\nTop 10 results by absolute difference:")
        print(results_df.reindex(
            results_df["difference"].abs().sort_values(ascending=False).index
        ).head(10).to_string())

    return results_df

## Plot the above results in a heatmap ##
def plot_region_pair_heatmap(results_df):
    """
    Plot heatmap of interaction type differences between cognate and non-cognate
    for each VH-VL region pair.
    """
    import matplotlib.colors as mcolors

    # Filter to most common interaction types based on mean count across all region pairs
    interaction_means = results_df.groupby("interaction_type")["cognate_mean"].mean().sort_values(ascending=False)
    common_interactions = interaction_means[interaction_means > 0.1].index.tolist()
    print(f"Keeping {len(common_interactions)} interaction types with mean count > 0.1:")
    print(common_interactions)

    # Filter to common interaction types only
    plot_df = results_df[results_df["interaction_type"].isin(common_interactions)].copy()

    # Filter to region pairs that have at least some contacts
    active_pairs = plot_df.groupby("region_pair").apply(
        lambda x: (x["cognate_mean"].abs().sum() + x["noncognate_mean"].abs().sum()) > 0
    )
    active_pairs = active_pairs[active_pairs].index
    plot_df = plot_df[plot_df["region_pair"].isin(active_pairs)]

    # Pivot to matrix: rows = region pairs, columns = interaction types
    diff_matrix = plot_df.pivot(index="region_pair", columns="interaction_type", values="difference")
    sig_matrix = plot_df.pivot(index="region_pair", columns="interaction_type", values="significant")

    # Reorder columns by mean count (most common first)
    cols_ordered = [c for c in common_interactions if c in diff_matrix.columns]
    diff_matrix = diff_matrix[cols_ordered]
    sig_matrix = sig_matrix[cols_ordered]

    # Sort rows by total absolute difference
    row_order = diff_matrix.abs().sum(axis=1).sort_values(ascending=False).index
    diff_matrix = diff_matrix.loc[row_order]
    sig_matrix = sig_matrix.loc[row_order]

    n_rows, n_cols = diff_matrix.shape
    fig, ax = plt.subplots(figsize=(n_cols * 1.2 + 2, n_rows * 0.6 + 2))

    # Diverging colormap - blue for more in cognate, red for more in non-cognate
    vmax = diff_matrix.abs().max().max()
    vmin = -vmax
    cmap = mcolors.LinearSegmentedColormap.from_list(
        "cog_noncog", ["#922B21", "#FFFFFF", "#1A5276"]
    )

    im = ax.imshow(diff_matrix.values, cmap=cmap, vmin=vmin, vmax=vmax, aspect='auto')

    # Add star for significant cells
    for i in range(n_rows):
        for j in range(n_cols):
            if sig_matrix.iloc[i, j]:
                ax.text(j, i, '*', ha='center', va='center',
                       fontsize=14, fontweight='bold', color='black')

    # Axis labels
    ax.set_xticks(range(n_cols))
    ax.set_xticklabels(diff_matrix.columns, rotation=45, ha='right', fontsize=10)
    ax.set_yticks(range(n_rows))
    ax.set_yticklabels(diff_matrix.index, fontsize=10)

    ax.set_xlabel("Interaction type", fontsize=11)
    ax.set_ylabel("VH-VL region pair", fontsize=11)
    ax.set_title("Interaction type differences per VH-VL region pair\n"
                "Cognate - Non-cognate (blue = more in cognate, red = more in non-cognate)\n"
                "* p_adj < 0.05", fontsize=11)

    # Colorbar
    cbar = plt.colorbar(im, ax=ax, shrink=0.5, pad=0.02)
    cbar.set_label("Mean count difference\n(cognate - non-cognate)", fontsize=9)

    plt.tight_layout()
    plt.savefig("region_pair_interaction_heatmap.png", dpi=150, bbox_inches='tight')
    plt.close()
    print("Saved region_pair_interaction_heatmap.png")


def main():
    csv_path = "bispec_arpeggio_outputs/therasabdab_bispec_arpeggio_contacts.csv"
    bispec_interactions = pd.read_csv(csv_path)
    feature_matrix = build_feature_matrix(bispec_interactions, 27)
    
    cosine_path = "cosine_similarity_results.csv"
    if not os.path.exists(cosine_path):
        cosine_df = calculate_cosine_similarity(feature_matrix)
        cosine_df.to_csv(cosine_path, index=False)
    else:
        cosine_df = pd.read_csv(cosine_path)
    
    print(cosine_df.describe())
    plot_cosine_similarities(cosine_df)
    run_statistical_tests(cosine_df)

    # Feature comparison
    feature_diff_path = "feature_differences.csv"
    if not os.path.exists(feature_diff_path):
        results_df = compare_cognate_noncognate_features(feature_matrix)
        results_df.to_csv(feature_diff_path, index=False)
    else:
        results_df = pd.read_csv(feature_diff_path)
    
    print(f"\nNumber of significant features (p_adj < 0.05): {(results_df['p_adjusted'] < 0.05).sum()}")
    print("\nTop 10 most significant features:")
    print(results_df.head(10))
    plot_feature_differences(results_df)

    # Interaction count comparison
    counts_path = "interaction_counts.csv"
    if not os.path.exists(counts_path):
        interaction_results, counts_df = calculate_interaction_counts(bispec_interactions)
        interaction_results.to_csv(counts_path, index=False)
        counts_df.to_csv("per_structure_counts.csv", index=False)
    else:
        interaction_results = pd.read_csv(counts_path)
        counts_df = pd.read_csv("per_structure_counts.csv")

    plot_interaction_counts(interaction_results)

    print(f"\nNumber of significant interaction types (p_adj < 0.05): {(interaction_results['p_adjusted'] < 0.05).sum()}")
    print("\nInteraction count results:")
    print(interaction_results.to_string())

    # Load PCA components and UMAP embeddings
    principle_components = np.load("pca_components.npy")
    embedding = np.load("umap_embedding.npy")

    # PCA outlier investigation
    pca_comparison, pca_outliers, pca_embed_df = investigate_outliers(
        feature_matrix, counts_df, principle_components, eps=2.0, min_samples=5, method="pca")

    # UMAP outlier investigation  
    umap_comparison, umap_outliers, umap_embed_df = investigate_outliers(
        feature_matrix, counts_df, embedding, eps=1, min_samples=5, method="umap")

    # Contact frequency analysis (replicating Dudzic et al)
    contact_freq_path = "contact_frequencies.csv"
    if not os.path.exists(contact_freq_path):
        contact_freq, bgn_freq, end_freq = calculate_all_contact_frequencies(bispec_interactions)
        contact_freq.to_csv(contact_freq_path)
        bgn_freq.to_csv("bgn_position_frequencies.csv")
        end_freq.to_csv("end_position_frequencies.csv")
    else:
        contact_freq = pd.read_csv(contact_freq_path, index_col=0).squeeze()
        bgn_freq = pd.read_csv("bgn_position_frequencies.csv", index_col=[0,1]).squeeze()
        end_freq = pd.read_csv("end_position_frequencies.csv", index_col=[0,1]).squeeze()

    print("\nTop 20 most frequent contact position pairs:")
    print(contact_freq.head(20))

    # Check CDR3-FWR2 contacts specifically
    cdr3_positions = [str(x) for x in range(105, 118)]
    fwr2_positions = [str(x) for x in range(39, 56)]

    cdr3_fwr2_pairs = contact_freq[
        contact_freq.index.map(lambda x: 
            (x.split("_")[0][1:] in cdr3_positions and x.split("_")[1][1:] in fwr2_positions) or
            (x.split("_")[0][1:] in fwr2_positions and x.split("_")[1][1:] in cdr3_positions)
        )
    ]

    print(f"\nCDR3-FWR2 contact pairs found: {len(cdr3_fwr2_pairs)}")
    print(cdr3_fwr2_pairs)

    plot_contact_frequencies(contact_freq, bgn_freq, end_freq)

    # Position interaction tables
    for chain_id in ["H", "L"]:
        for cognate, label in [(True, "cognate"), (False, "noncognate")]:
            table_path = f"V{chain_id}_position_interactions_{label}.csv"
            if not os.path.exists(table_path):
                table = calculate_position_interaction_table(
                    bispec_interactions, chain_id, cognate)
                style_and_save_table(table, chain_id, label)
            else:
                table = pd.read_csv(table_path, index_col=0)
                print(f"\nV{chain_id} {label} (loaded from CSV):")
                print(table.to_string())

    # IMGT contact frequency plots (Dudzic Fig 4 style)
    for chain_id in ["H", "L"]:
        # Run statistical test once per chain (uses both cognate and non-cognate)
        sig_positions, position_stats = test_imgt_position_differences(bispec_interactions, chain_id)
        position_stats.to_csv(f"V{chain_id}_imgt_position_stats.csv", index=False)
        
        for cognate, label in [(True, "cognate"), (False, "noncognate")]:
            plot_imgt_contact_frequency(bispec_interactions, chain_id, cognate, label, sig_positions)

    # Plot slope chart
    plot_slope_chart_v2(bispec_interactions) 

    # Condensed slope chart
    plot_slope_chart_condensed(bispec_interactions, interval=10)

    # Differing pairs table
    differing_pairs = get_differing_pairs_table(bispec_interactions, threshold=0.1)   

    # ── Interaction types per IMGT region ────────────────────────────────────
    for chain_id in ["H", "L"]:
        region_interaction_results = test_interaction_types_per_region(
            bispec_interactions, chain_id)
        
    # ── Interaction types per VH-VL region pair ──────────────────────────────
    region_pair_results = test_interaction_types_per_region_pair(bispec_interactions)

    # ── Interaction types per VH-VL region pair ──────────────────────────────
    region_pair_stats_path = "region_pair_interaction_types.csv"
    if not os.path.exists(region_pair_stats_path):
        region_pair_results = test_interaction_types_per_region_pair(bispec_interactions)
    else:
        region_pair_results = pd.read_csv(region_pair_stats_path)
    
    plot_region_pair_heatmap(region_pair_results)

    feature_matrix.to_csv("feature_matrix.csv")


if __name__ == "__main__":
    main()